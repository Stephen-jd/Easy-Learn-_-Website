import datetime
import openpyxl
from io import BytesIO
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.http import HttpResponse, Http404
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Project, TrainerProfile, DailyWorkload, ClassWorkNote, Expense, Invoice

# ==============================================================================
# DECORATORS & ACCESS UTILITIES
# ==============================================================================

def trainer_access_required(view_func):
    """
    Ensures the user is logged in, is a trainer, has approved status, and is not deactivated.
    """
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        try:
            profile = request.user.trainer_profile
            if not profile.is_trainer_accessible:
                return render(request, 'core/deactivated.html', {
                    'profile': profile,
                    'is_deactivated_by_project': profile.is_deactivated_by_projects
                })
            return view_func(request, *args, **kwargs)
        except AttributeError:
            messages.error(request, "Access Denied: You are not registered as a Trainer.")
            return redirect('login')
    return wrapper

def admin_access_required(view_func):
    """
    Ensures the user is an admin (superuser).
    """
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, "Access Denied: Admin privileges required.")
            return redirect('landing')
        return view_func(request, *args, **kwargs)
    return wrapper

# ==============================================================================
# GENERAL & AUTHENTICATION VIEWS
# ==============================================================================

def landing_page(request):
    """
    Landing Page showing key operations stats for Easy Learn Academy.
    Admin gets full analytics operations landing page.
    Trainer is redirected to Trainer Dashboard.
    Anonymous user is shown a clean gateway login selector.
    """
    # 1. Access redirect for Trainer
    if request.user.is_authenticated:
        if not request.user.is_superuser:
            return redirect('trainer_dashboard')
    else:
        # 2. Show gateway portal to anonymous guests
        return render(request, 'core/landing_gate.html')

    # 3. Operations dashboard shown only to Admin
    stats = {
        'total_active_projects': Project.objects.filter(status='Active').count(),
        'total_completed_projects': Project.objects.filter(status='Completed').count(),
        'active_trainers_count': TrainerProfile.objects.filter(is_active=True, is_approved=True).count(),
        'old_trainers_count': TrainerProfile.objects.filter(Q(is_active=False) | Q(projects__status='Completed')).distinct().count(),
        'total_expenses': Expense.objects.aggregate(total=Sum('amount'))['total'] or 0.0,
    }
    
    # List of current ongoing projects
    current_projects = Project.objects.filter(status='Active').order_by('start_date')
    
    # List of active onboarding trainers
    active_trainers = TrainerProfile.objects.filter(is_active=True, is_approved=True)

    # Vendor Details
    vendors = [
        {"name": "Easy Learn Solutions", "contact": "manager.projects@easy-learn-academy.com", "service": "Core Edtech Content & Infrastructure"},
        {"name": "Udupi Tech Ventures", "contact": "info@udupitech.in", "service": "Logistics & Offline Lab Support"},
        {"name": "Vijayapur Edu Partner", "contact": "admin@vjedu.org", "service": "Local Coordinator & Venue Vendor"}
    ]

    context = {
        'stats': stats,
        'current_projects': current_projects,
        'active_trainers': active_trainers,
        'vendors': vendors,
    }
    return render(request, 'core/landing.html', context)

def login_view(request):
    """
    Consolidated Login page for Trainer and Admin.
    Enforces that trainers must be approved by the admin before they can log in.
    """
    if request.user.is_authenticated:
        return redirect('landing' if not request.user.is_superuser else 'admin_dashboard')
        
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        if user is not None:
            if user.is_superuser:
                login(request, user)
                messages.success(request, f"Welcome back, Admin!")
                return redirect('admin_dashboard')
            else:
                try:
                    profile = user.trainer_profile
                    if not profile.is_approved:
                        messages.error(request, "Login Denied: Your partner trainer profile is currently pending administrator approval. You will be able to log in once approved.")
                        return redirect('login')
                    elif not profile.is_active:
                        messages.error(request, "Login Denied: Your trainer profile has been administrative deactivated.")
                        return redirect('login')
                    else:
                        login(request, user)
                        messages.success(request, f"Welcome back, {user.first_name}!")
                        return redirect('trainer_dashboard')
                except AttributeError:
                    messages.error(request, "Login Failed: Trainer profile not found.")
                    return redirect('login')
        else:
            messages.error(request, "Invalid username or password.")
            
    return render(request, 'core/login.html')

def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('landing')

# ==============================================================================
# TRAINER VIEWS & REGISTRATION
# ==============================================================================

def trainer_signup(request):
    """
    Trainer signup allowing doc upload (Photo, Aadhaar file), quoted rates, bank details, and project assignment.
    """
    if request.user.is_authenticated:
        return redirect('landing')
        
    projects = Project.objects.filter(status='Active')
    
    if request.method == 'POST':
        # User details
        uname = request.POST.get('username')
        email = request.POST.get('email')
        pwd = request.POST.get('password')
        fname = request.POST.get('first_name')
        lname = request.POST.get('last_name')
        
        # Profile details
        aadhaar_no = request.POST.get('aadhaar_number')
        pan = request.POST.get('pan_card')
        quote_amt = request.POST.get('quoted_amount')
        quote_type = request.POST.get('quoted_type')
        
        # Files
        photo_file = request.FILES.get('photo')
        aadhaar_file = request.FILES.get('aadhaar_file')
        
        # Bank info
        bname = request.POST.get('bank_name')
        bacc = request.POST.get('bank_account_number')
        bifsc = request.POST.get('bank_ifsc')
        upi = request.POST.get('upi_id')
        
        # Project selection
        proj_id = request.POST.get('project')

        # Validation
        if User.objects.filter(username=uname).exists():
            messages.error(request, "Username already exists. Please choose a unique one.")
            return render(request, 'core/trainer_signup.html', {'projects': projects})
            
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already registered.")
            return render(request, 'core/trainer_signup.html', {'projects': projects})

        try:
            with transaction.atomic():
                # Create user
                new_user = User.objects.create_user(
                    username=uname,
                    email=email,
                    password=pwd,
                    first_name=fname,
                    last_name=lname
                )
                
                # Create trainer profile (is_approved=False by default)
                profile = TrainerProfile.objects.create(
                    user=new_user,
                    photo=photo_file,
                    aadhaar_number=aadhaar_no,
                    aadhaar_file=aadhaar_file,
                    pan_card=pan,
                    bank_name=bname,
                    bank_account_number=bacc,
                    bank_ifsc=bifsc,
                    upi_id=upi,
                    quoted_amount=quote_amt,
                    quoted_type=quote_type,
                    is_approved=False,  # Needs manual Admin approval
                    is_active=True
                )
                
                if proj_id:
                    p = Project.objects.get(id=proj_id)
                    profile.projects.add(p)
                
                messages.success(request, "Registration successful! Your partner trainer application is currently pending review. Once the administrator approves your profile, you will be able to log in to your dashboard.")
                return redirect('login')
                
        except Exception as e:
            messages.error(request, f"Error during signup: {str(e)}")
            
    return render(request, 'core/trainer_signup.html', {'projects': projects})

@trainer_access_required
def trainer_dashboard(request):
    """
    Main Trainer Dashboard for managing profile details, logs, daily updates, and raising invoices.
    """
    profile = request.user.trainer_profile
    my_projects = profile.projects.all()
    my_workloads = DailyWorkload.objects.filter(trainer=request.user).order_by('-date')
    my_invoices = Invoice.objects.filter(trainer=request.user).order_by('-created_at')
    
    # Calculate sum of sessions logged
    total_sessions = my_workloads.aggregate(Sum('sessions_count'))['sessions_count__sum'] or 0

    context = {
        'profile': profile,
        'projects': my_projects,
        'workloads': my_workloads,
        'invoices': my_invoices,
        'total_sessions': total_sessions,
    }
    return render(request, 'core/trainer_dashboard.html', context)

@trainer_access_required
def update_trainer_profile(request):
    """
    Allows the trainer to modify their photo, bank details, PAN card, Aadhaar, and UPI.
    """
    profile = request.user.trainer_profile
    if request.method == 'POST':
        # Update details
        request.user.first_name = request.POST.get('first_name')
        request.user.last_name = request.POST.get('last_name')
        request.user.save()

        profile.bank_name = request.POST.get('bank_name')
        profile.bank_account_number = request.POST.get('bank_account_number')
        profile.bank_ifsc = request.POST.get('bank_ifsc')
        profile.upi_id = request.POST.get('upi_id')
        profile.pan_card = request.POST.get('pan_card')
        profile.aadhaar_number = request.POST.get('aadhaar_number')
        
        # File uploads if supplied
        if request.FILES.get('photo'):
            profile.photo = request.FILES.get('photo')
        if request.FILES.get('aadhaar_file'):
            profile.aadhaar_file = request.FILES.get('aadhaar_file')
            
        profile.save()
        messages.success(request, "Personal & Bank profile details updated successfully!")
        return redirect('trainer_dashboard')
    
    return redirect('trainer_dashboard')

@trainer_access_required
def add_workload(request):
    """
    Log daily session workload with multiple class notes in PDF format.
    """
    if request.method == 'POST':
        proj_id = request.POST.get('project')
        log_date_str = request.POST.get('date')
        sessions = int(request.POST.get('sessions_count'))
        topics = request.POST.get('topics_covered')
        subtopics = request.POST.get('subtopics_covered')
        files = request.FILES.getlist('notes_files')
        
        # Convert date
        log_date = datetime.datetime.strptime(log_date_str, "%Y-%m-%d").date()
        
        # Verify trainer is assigned to this project
        project = get_object_or_404(Project, id=proj_id)
        if project not in request.user.trainer_profile.projects.all():
            messages.error(request, "Access Denied: You are not assigned to this project.")
            return redirect('trainer_dashboard')
            
        # Verify notes files are PDF only
        for f in files:
            if not f.name.lower().endswith('.pdf'):
                messages.error(request, f"Upload Failed: '{f.name}' is not in PDF format. Only PDF files are allowed for class notes.")
                return redirect('trainer_dashboard')

        # Auto-compute joint summaries for backward compatibility
        computed_topics = "; ".join([request.POST.get(f'session_content_{i}', '') for i in range(1, sessions + 1)])
        computed_subtopics = "; ".join([
            f"Session {i} [{request.POST.get(f'session_time_{i}', '')} - {request.POST.get(f'session_batch_{i}', '')}]: {request.POST.get(f'session_content_{i}', '')[:45]}..."
            for i in range(1, sessions + 1)
        ])

        try:
            with transaction.atomic():
                # Save daily workload log
                workload, created = DailyWorkload.objects.get_or_create(
                    trainer=request.user,
                    project=project,
                    date=log_date,
                    defaults={
                        'sessions_count': sessions,
                        'topics_covered': computed_topics,
                        'subtopics_covered': computed_subtopics
                    }
                )
                
                if not created:
                    # Update fields if it already exists
                    workload.sessions_count = sessions
                    workload.topics_covered = computed_topics
                    workload.subtopics_covered = computed_subtopics
                    workload.save()

                # Wipe old session details for clean overwrite
                workload.session_details.all().delete()

                # Process and save SessionDetail records
                for i in range(1, sessions + 1):
                    time_val = request.POST.get(f'session_time_{i}', '')
                    batch_val = request.POST.get(f'session_batch_{i}', '')
                    content_val = request.POST.get(f'session_content_{i}', '')
                    
                    SessionDetail.objects.create(
                        workload=workload,
                        session_num=i,
                        time_slot=time_val,
                        batch=batch_val,
                        content_covered=content_val
                    )

                # Process PDF notes files
                for f in files:
                    ClassWorkNote.objects.create(workload=workload, file=f)
                    
            messages.success(request, f"Daily workload with {sessions} individual sessions logged successfully for {log_date}!")
        except Exception as e:
            messages.error(request, f"Error logging workload: {str(e)}")
            
    return redirect('trainer_dashboard')

@trainer_access_required
def raise_invoice(request):
    """
    Generate trainer invoice. Automatically calculates total working days (based on logged days),
    deducts money for sick/medical leaves, and submits to admin profile.
    """
    if request.method == 'POST':
        proj_id = request.POST.get('project')
        month_str = request.POST.get('month') # YYYY-MM
        holidays = int(request.POST.get('total_holidays') or 0)
        leaves = int(request.POST.get('sick_leaves') or 0)
        
        project = get_object_or_404(Project, id=proj_id)
        profile = request.user.trainer_profile
        
        # Get start/end dates for selected month to count working days
        try:
            year, month = map(int, month_str.split('-'))
        except ValueError:
            messages.error(request, "Invalid month format. Please select a valid month.")
            return redirect('trainer_dashboard')
            
        # Count daily workloads logged by this trainer for this project in the selected month
        workload_days = DailyWorkload.objects.filter(
            trainer=request.user,
            project=project,
            date__year=year,
            date__month=month
        ).count()
        
        if workload_days == 0:
            messages.warning(request, f"You haven't logged any daily workload entries for {month_str}. Please log your sessions first.")
            return redirect('trainer_dashboard')

        # Check if invoice already exists
        if Invoice.objects.filter(trainer=request.user, project=project, month=month_str).exists():
            messages.warning(request, f"An invoice has already been generated/raised for {month_str} for project {project.name}.")
            return redirect('trainer_dashboard')

        # Compute Quote amounts & Deductions
        rate = profile.quoted_amount
        rate_type = profile.quoted_type
        
        # If rate is Daily:
        if rate_type == 'Daily':
            working_days = workload_days
            # Deduction is simply leaves * daily rate
            deductions = leaves * rate
            subtotal = working_days * rate
            total = max(0, subtotal - deductions)
        else: # Monthly
            # Assume 30 calendar days or get daily equivalent as Monthly / 24 standard working days
            standard_working_days = 24
            daily_equivalent_rate = rate / standard_working_days
            working_days = workload_days
            
            # Subtotal could be the monthly rate scaled by working days vs standard 24, or standard full monthly rate
            # Let's say: if they worked the full logged month, they get the quoted amount, but sick leaves deduct from it.
            deductions = leaves * daily_equivalent_rate
            subtotal = rate
            total = max(0, subtotal - deductions)

        # Create Invoice
        invoice = Invoice.objects.create(
            trainer=request.user,
            project=project,
            month=month_str,
            total_working_days=working_days,
            total_holidays=holidays,
            sick_leaves=leaves,
            deductions=deductions,
            total_amount=total,
            status='Pending'
        )
        
        messages.success(request, f"Invoice raised successfully for {month_str}! Net amount calculated: ₹{total:.2f}. Status is Pending Review.")
        
    return redirect('trainer_dashboard')

@login_required
def download_invoice_pdf(request, invoice_id):
    """
    Generates a beautifully styled dynamic PDF invoice matching the logo theme (Orange & Blue).
    Available to the raising trainer and Admin.
    """
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # Authorization check
    if not request.user.is_superuser and invoice.trainer != request.user:
        raise Http404("Access Denied: You do not have permission to view this invoice.")

    # Prepare context
    trainer_profile = get_object_or_404(TrainerProfile, user=invoice.trainer)
    
    context = {
        'invoice': invoice,
        'profile': trainer_profile,
        'date_generated': datetime.date.today(),
        'due_date': datetime.date.today() + datetime.timedelta(days=15)
    }

    # Render PDF template to response
    template = get_template('core/invoice_pdf.html')
    html = template.render(context)
    result = BytesIO()
    
    # xhtml2pdf generation
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f"Invoice_{invoice.trainer.username}_{invoice.project.name.replace(' ', '_')}_{invoice.month}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    return HttpResponse("Error generating PDF invoice", status=500)

# ==============================================================================
# ADMIN VIEW & CONTROLS
# ==============================================================================

@admin_access_required
def admin_dashboard(request):
    """
    Tabbed Operations Panel for administrators.
    Tabs:
    1. Project Manager: Create, Edit, Toggle Complete.
    2. Trainer Manager: Credentials, View single Trainer details, approval, dates, Covered topics date-range report.
    3. Expense Tracker: Log project expense, Monthly analytics dashboards, Excel export.
    4. Invoice Manager: Monthly total receiver board, filter project, Approve/Pay.
    """
    # 1. Projects
    projects = Project.objects.all().order_by('-created_at')
    active_projects = Project.objects.filter(status='Active')
    
    # 2. Trainers
    trainers = TrainerProfile.objects.all().order_by('-user__date_joined')
    
    # 3. Expenses
    expenses = Expense.objects.all().order_by('-date')
    expense_categories = [cat[0] for cat in Expense.CATEGORY_CHOICES]
    
    # Total monthly expense summary
    monthly_expenses = Expense.objects.values('date__year', 'date__month').annotate(total=Sum('amount')).order_by('-date__year', '-date__month')
    
    # Beautiful chart datasets computed via standard Aggregates
    cat_distribution = Expense.objects.values('category').annotate(total=Sum('amount'))
    chart_labels = [c['category'] for c in cat_distribution]
    chart_values = [float(c['total']) for c in cat_distribution]
    
    # Project distribution
    proj_distribution = Expense.objects.values('project__name').annotate(total=Sum('amount'))
    proj_labels = [p['project__name'] for p in proj_distribution]
    proj_values = [float(p['total']) for p in proj_distribution]

    # 4. Invoices
    invoices = Invoice.objects.all().order_by('-created_at')
    
    # Group invoices by month for invoice navbar view
    monthly_invoice_sums = Invoice.objects.values('month').annotate(total=Sum('total_amount'), count=Count('id')).order_by('-month')

    context = {
        'projects': projects,
        'active_projects': active_projects,
        'trainers': trainers,
        'expenses': expenses,
        'expense_categories': expense_categories,
        'monthly_expenses': monthly_expenses,
        'chart_labels': chart_labels,
        'chart_values': chart_values,
        'proj_labels': proj_labels,
        'proj_values': proj_values,
        'invoices': invoices,
        'monthly_invoice_sums': monthly_invoice_sums,
        'is_admin': True,
    }
    return render(request, 'core/admin_dashboard.html', context)

@admin_access_required
def admin_view_trainer(request, trainer_id):
    """
    Detailed trainer review dashboard: bank credentials, PAN, Aadhaar, quoted rates, active projects,
    and a table displaying daily topics covered throughout dates with filters.
    """
    trainer_profile = get_object_or_404(TrainerProfile, id=trainer_id)
    trainer_user = trainer_profile.user
    
    # Filter projects assigned
    projects = trainer_profile.projects.all()
    
    # Filter workloads
    workloads = DailyWorkload.objects.filter(trainer=trainer_user).order_by('-date')
    
    context = {
        'profile': trainer_profile,
        'trainer_user': trainer_user,
        'projects': projects,
        'workloads': workloads,
    }
    return render(request, 'core/admin_view_trainer.html', context)

@admin_access_required
def admin_toggle_trainer_status(request, trainer_id):
    """
    Allows admin to activate or deactivate the trainer's status.
    """
    trainer_profile = get_object_or_404(TrainerProfile, id=trainer_id)
    trainer_profile.is_active = not trainer_profile.is_active
    trainer_profile.save()
    
    status_label = "Activated" if trainer_profile.is_active else "Deactivated"
    messages.success(request, f"Trainer {trainer_profile.user.first_name} has been successfully {status_label}.")
    return redirect('admin_view_trainer', trainer_id=trainer_profile.id)

@admin_access_required
def admin_add_project(request):
    """
    Add a new training Project college.
    """
    if request.method == 'POST':
        name = request.POST.get('name')
        duration = int(request.POST.get('duration_days') or 0)
        sdate = request.POST.get('start_date')
        edate = request.POST.get('end_date')
        tech = request.POST.get('technology')
        sem = request.POST.get('semester_year')
        
        Project.objects.create(
            name=name,
            duration_days=duration,
            start_date=sdate,
            end_date=edate,
            technology=tech,
            semester_year=sem,
            status='Active'
        )
        messages.success(request, f"Project college '{name}' added successfully!")
    return redirect('admin_dashboard')

@admin_access_required
def admin_edit_project(request, project_id):
    """
    Modify details of an existing college project.
    """
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        project.name = request.POST.get('name')
        project.duration_days = int(request.POST.get('duration_days') or 0)
        project.start_date = request.POST.get('start_date')
        project.end_date = request.POST.get('end_date')
        project.technology = request.POST.get('technology')
        project.semester_year = request.POST.get('semester_year')
        project.save()
        messages.success(request, f"Project '{project.name}' details updated successfully!")
    return redirect('admin_dashboard')

@admin_access_required
def admin_toggle_project_status(request, project_id):
    """
    Updates the status of a project. When set to Completed:
    Trainer accounts assigned ONLY to completed projects will be automatically locked out/deactivated.
    """
    project = get_object_or_404(Project, id=project_id)
    project.status = 'Completed' if project.status == 'Active' else 'Active'
    project.save()
    
    status_label = "Completed" if project.status == 'Completed' else "Reactivated (Active)"
    messages.success(request, f"Project '{project.name}' status set to {status_label}.")
    
    # Auto deactivation/reactivation updates
    # We will trigger a loop to re-check all trainers assigned to this project
    for trainer in project.trainers.all():
        if trainer.is_deactivated_by_projects:
            messages.info(request, f"Trainer '{trainer.user.first_name}' has been automatically locked/deactivated since all their projects are completed.")
            
    return redirect('admin_dashboard')

@admin_access_required
def admin_add_expense(request):
    """
    Record expense against project and month.
    """
    if request.method == 'POST':
        proj_id = request.POST.get('project')
        exp_date = request.POST.get('date')
        cat = request.POST.get('category')
        amt = float(request.POST.get('amount') or 0.0)
        desc = request.POST.get('description')
        
        project = get_object_or_404(Project, id=proj_id)
        Expense.objects.create(
            project=project,
            date=exp_date,
            category=cat,
            amount=amt,
            description=desc
        )
        messages.success(request, f"Expense of ₹{amt:.2f} logged successfully for {project.name}!")
        
    return redirect('admin_dashboard')

@admin_access_required
def admin_update_invoice_status(request, invoice_id):
    """
    Approve invoice or mark as Paid.
    """
    invoice = get_object_or_404(Invoice, id=invoice_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in ['Approved', 'Paid', 'Pending']:
            invoice.status = new_status
            invoice.save()
            messages.success(request, f"Invoice status for {invoice.trainer.username} updated to '{new_status}' successfully!")
    return redirect('admin_dashboard')

# ==============================================================================
# REPORT EXPORTING & DOCUMENT DOWNLOADS
# ==============================================================================

@admin_access_required
def admin_download_project_report(request, trainer_id, project_id):
    """
    Downloads a detailed project report (PDF) for a trainer showing covering dates,
    topics covered throughout daily session tables between a start date and custom end date.
    """
    trainer_profile = get_object_or_404(TrainerProfile, id=trainer_id)
    project = get_object_or_404(Project, id=project_id)
    
    # Custom dates from GET
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    workloads = DailyWorkload.objects.filter(trainer=trainer_profile.user, project=project)
    
    if start_date_str:
        workloads = workloads.filter(date__gte=start_date_str)
    if end_date_str:
        workloads = workloads.filter(date__lte=end_date_str)
        
    workloads = workloads.order_by('date')
    
    context = {
        'profile': trainer_profile,
        'project': project,
        'workloads': workloads,
        'start_date': start_date_str or project.start_date,
        'end_date': end_date_str or datetime.date.today(),
        'date_generated': datetime.date.today()
    }
    
    template = get_template('core/project_report_pdf.html')
    html = template.render(context)
    result = BytesIO()
    
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f"Report_{trainer_profile.user.username}_{project.name.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    return HttpResponse("Error generating project report PDF", status=500)

@admin_access_required
def admin_download_expenses_excel(request):
    """
    Generates a professionally styled Excel file containing all logged expenses,
    sorted by project and month. Uses openpyxl to deliver clean typography & colors.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Academy Expenses"
    
    # Styles
    title_font = Font(name='Segoe UI', size=16, bold=True, color='0747A6')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Segoe UI', size=10)
    total_font = Font(name='Segoe UI', size=11, bold=True)
    
    fill_header = PatternFill(start_color='0747A6', end_color='0747A6', fill_type='solid')
    fill_total = PatternFill(start_color='FFEBD6', end_color='FFEBD6', fill_type='solid') # soft orange
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Title Block
    ws.merge_cells('A1:E1')
    ws['A1'] = "Easy Learn Academy - Expense Tracker Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = align_left
    ws.row_dimensions[1].height = 30
    
    # Date Generated
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Generated: {datetime.date.today().strftime('%B %d, %Y')}"
    ws['A2'].font = Font(name='Segoe UI', size=10, italic=True)
    ws['A2'].alignment = align_left
    
    # Headers
    headers = ["Project College", "Date Logged", "Expense Category", "Description/Details", "Amount (INR)"]
    ws.append([]) # Blank row
    ws.append(headers)
    
    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 25
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_num)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = align_left if col_num != 2 and col_num != 5 else align_center
        cell.border = thin_border
        
    # Data Rows
    expenses = Expense.objects.all().order_by('project__name', '-date')
    row_idx = 5
    total_sum = 0
    
    for exp in expenses:
        ws.row_dimensions[row_idx].height = 20
        
        c1 = ws.cell(row=row_idx, column=1, value=exp.project.name)
        c2 = ws.cell(row=row_idx, column=2, value=exp.date.strftime('%Y-%m-%d'))
        c3 = ws.cell(row=row_idx, column=3, value=exp.category)
        c4 = ws.cell(row=row_idx, column=4, value=exp.description or "-")
        c5 = ws.cell(row=row_idx, column=5, value=float(exp.amount))
        
        c1.alignment = align_left
        c2.alignment = align_center
        c3.alignment = align_left
        c4.alignment = align_left
        c5.alignment = align_right
        c5.number_format = '₹#,##0.00'
        
        for c in [c1, c2, c3, c4, c5]:
            c.font = data_font
            c.border = thin_border
            
        total_sum += exp.amount
        row_idx += 1
        
    # Total Summary Row
    ws.row_dimensions[row_idx].height = 24
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
    total_label_cell = ws.cell(row=row_idx, column=1, value="Grand Total Spending")
    total_label_cell.font = total_font
    total_label_cell.alignment = align_right
    total_label_cell.fill = fill_total
    
    total_val_cell = ws.cell(row=row_idx, column=5, value=float(total_sum))
    total_val_cell.font = total_font
    total_val_cell.alignment = align_right
    total_val_cell.fill = fill_total
    total_val_cell.number_format = '₹#,##0.00'
    
    for col in range(1, 6):
        ws.cell(row=row_idx, column=col).border = thin_border
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            # Skip title row and merged cell formulas
            if cell.row in [1, 2] or (cell.row == row_idx and cell.column < 5):
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Save Workbook to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Easy_Learn_Academy_Expenses.xlsx"'
    wb.save(response)
    
    return response
